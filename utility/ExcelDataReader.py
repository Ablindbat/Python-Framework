import openpyxl


class Excel:
    # path of excel file
    path = "..//TestData//TestData.xlsx"
    # dictionary variable
    excel_Dict = {}

    # excel reader function
    def read_excel(self, data):
        book = openpyxl.load_workbook(self.path)
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == data:
                for j in range(2, sheet.max_column + 1):
                    self.excel_Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                break
        return self.excel_Dict

